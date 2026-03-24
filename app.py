import os
import json
import hashlib
import hmac
import base64
import logging
import re
import threading
import time as time_module
from datetime import datetime, timedelta, timezone
from http.server import HTTPServer, BaseHTTPRequestHandler
from io import BytesIO
from urllib.parse import unquote
from collections import OrderedDict

import anthropic
from google.oauth2 import service_account
from googleapiclient.discovery import build
import requests
from openpyxl import Workbook
from openpyxl.styles import Font

LINE_CHANNEL_SECRET = os.environ.get("LINE_CHANNEL_SECRET", "")
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
GOOGLE_CALENDAR_ID = os.environ.get("GOOGLE_CALENDAR_ID", "")
GOOGLE_CREDENTIALS_PATH = "/app/credentials.json"
PORT = int(os.environ.get("PORT", "8000"))
REPORT_HOUR = int(os.environ.get("REPORT_HOUR", "18"))
ADMIN_USER_ID = os.environ.get("ADMIN_USER_ID", "")
NAS_EXTERNAL_URL = os.environ.get("NAS_EXTERNAL_URL", "")
REPORT_DIR = os.environ.get("REPORT_DIR", "/app/reports")

os.makedirs(REPORT_DIR, exist_ok=True)

TW_TZ = timezone(timedelta(hours=8))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

pending_confirmations = {}
daily_event_log = []
daily_event_log_date = None

def get_calendar_service():
    credentials = service_account.Credentials.from_service_account_file(
        GOOGLE_CREDENTIALS_PATH, scopes=["https://www.googleapis.com/auth/calendar"])
    return build("calendar", "v3", credentials=credentials)

def parse_event_with_claude(message_text):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    now = datetime.now(TW_TZ)
    today = now.strftime("%Y-%m-%d")
    current_year = now.strftime("%Y")
    prompt = f"""你是一個行程解析助手。請從以下 Line 群組訊息中擷取行程資訊。
今天的日期是 {today}。

請先判斷這是「一般行程」還是「同仁請假」。
判斷方式：如果訊息中包含「請假」、「休假」、「特休」、「病假」、「事假」、「補休」、「產假」、「產檢假」、「喪假」等關鍵字，就是「同仁請假」類型。

=== 如果是「一般行程」，回傳以下 JSON ===
{{
    "type": "event",
    "title": "行程/會議名稱",
    "date": "YYYY-MM-DD",
    "end_date": "YYYY-MM-DD",
    "start_time": "HH:MM",
    "end_time": "HH:MM",
    "location": "完整地點",
    "location_city": "縣市簡稱",
    "host": "主持人姓名",
    "attendees": ["出席人員1", "出席人員2"],
    "staff": "會議幕僚姓名",
    "meeting_url": "線上會議連結",
    "meeting_id": "會議號",
    "meeting_password": "會議密碼",
    "notes": "其他備註資訊"
}}

=== 如果是「同仁請假」，回傳以下 JSON ===
{{
    "type": "leave",
    "leaves": [
        {{"name": "同仁姓名", "date": "YYYY-MM-DD", "leave_type": "假別", "start_time": "HH:MM", "end_time": "HH:MM"}}
    ]
}}

注意：每筆 leave 記錄都必須包含自己的 date 欄位，表示這筆請假是哪一天的。
如果同一位同仁跨多天請假，每天要各自一筆 leave，date 各自不同。
如果同一位同仁同一天有多個不同時段，也要分成多筆 leave（date 相同但時間不同）。

例如：
「郁穎3/25 14:00-17:00、3/26 08:00-17:00、3/27 08:00-17:00」應拆成三筆：
{{"name": "郁穎", "date": "2026-03-25", "leave_type": null, "start_time": "14:00", "end_time": "17:00"}}
{{"name": "郁穎", "date": "2026-03-26", "leave_type": null, "start_time": "08:00", "end_time": "17:00"}}
{{"name": "郁穎", "date": "2026-03-27", "leave_type": null, "start_time": "08:00", "end_time": "17:00"}}

「張三3/10 0900-1200產檢假、1400-1800請休」應拆成兩筆：
{{"name": "張三", "date": "2026-03-10", "leave_type": "產檢假", "start_time": "09:00", "end_time": "12:00"}}
{{"name": "張三", "date": "2026-03-10", "leave_type": null, "start_time": "14:00", "end_time": "18:00"}}

規則：
1. 如果沒有提到年份，預設使用 {current_year} 年
2. 如果沒有提到結束時間，且有開始時間，一般行程預設為開始時間後 1 小時
3. 如果某個欄位訊息中沒有提到，填入 null
4. 日期請務必轉換成 YYYY-MM-DD（西元年）格式
5. 時間請務必轉換成 24 小時制 HH:MM 格式
6. 「下週一」「這週五」等相對日期請根據今天日期計算出確切日期
7. 只回傳 JSON，不要有任何其他文字
8. 線上會議資訊：如果訊息中有會議連結、會議號、密碼等資訊，請分別擷取
9. 會議號和密碼可能用「/」分隔（例如「2510 771 7063/2026」表示會議號為「2510 771 7063」，密碼為「2026」）
10. 主持人可能以「主持人」、「主席」、「召集人」等詞彙標示。注意：「講師」不是主持人，講師資訊請放在 notes 欄位
    - 「工作人員」、「幕僚」等詞彙代表工作人員，請填入 staff 欄位
    - staff 欄位也可以用來記錄負責的辦公室或單位（例如「台北辦公室」）
11. 請假中的「0213」代表 2 月 13 日，「0315」代表 3 月 15 日
12. 請假時間如「(09-1030)」代表 09:00 到 10:30
13. 一則請假訊息中可能包含多位同仁的請假資訊
14. 出席人員格式規則：
    - 人名後面的(O)代表會出席，(X)代表不出席，請保留這個標記，括號用半形
    - 「幕僚」後面的人名代表會議幕僚，請格式化為「幕僚(人名)」，括號用半形
    - 多位人員用頓號「、」分隔
    - 例如原文「朱理事長(O) 幕僚 貝珊」應整理成 attendees: ["朱理事長(O)", "幕僚(貝珊)"]，staff: "貝珊"
15. location_city 只填縣市簡稱：台北市→台北、新北市→新北、桃園市→桃園、台中市→台中、台南市→台南、高雄市→高雄，以此類推。如果地點是線上或無法判斷縣市則填 null
16. 請假假別（leave_type）辨識規則：
    - 如果訊息有註明假別，leave_type 填寫對應假別：「產假」、「產檢假」、「病假」、「喪假」、「事假」、「補休」、「特休」、「公假」、「婚假」等
    - 如果訊息只寫「請假」或「休假」而沒有具體假別，leave_type 填 null
    - 假別可能出現在人名前面或後面，例如「張三產檢假1小時」或「產檢假 張三1小時」
17. 民國年轉換：「115年」=西元2026年、「114年」=西元2025年，以此類推（民國年+1911=西元年）
18. 跨天行程：
    - 如果行程跨多天（例如「8月13日至8月15日」），date 填開始日期，end_date 填結束日期
    - 如果行程只有一天，end_date 填 null
    - 如果沒有明確的開始和結束時間（只有日期），start_time 和 end_time 填 null（代表全天事件）

訊息內容：
{message_text}"""
    response = client.messages.create(model="claude-sonnet-4-20250514", max_tokens=1024, messages=[{"role": "user", "content": prompt}])
    response_text = response.content[0].text.strip()
    if response_text.startswith("```"):
        response_text = response_text.split("\n", 1)[1]
        response_text = response_text.rsplit("```", 1)[0].strip()
    return json.loads(response_text)

def normalize_str(s):
    """正規化字串：移除所有空白，用於模糊比對"""
    return re.sub(r"\s+", "", s).strip()

def find_duplicate_event(service, event_data):
    date_str = event_data["date"]
    events_result = service.events().list(
        calendarId=GOOGLE_CALENDAR_ID, timeMin=f"{date_str}T00:00:00+08:00",
        timeMax=f"{date_str}T23:59:59+08:00", singleEvents=True, orderBy="startTime").execute()
    new_title = normalize_str(event_data.get("title") or "")
    new_location = normalize_str(event_data.get("location") or "")
    for evt in events_result.get("items", []):
        if evt.get("summary") == "同仁休假登記":
            continue
        evt_title = normalize_str(evt.get("summary") or "")
        evt_location = normalize_str(evt.get("location") or "")
        if (new_title and evt_title and new_title == evt_title) or (new_location and evt_location and new_location == evt_location):
            return evt
    return None

def build_event_body(event_data, message_timestamp=None):
    parts = []
    if event_data.get("host"):
        parts.append(f"🎤 主持人：{event_data['host']}")
    if event_data.get("attendees"):
        parts.append(f"📋 出席人員：{'、'.join(event_data['attendees'])}")
    if event_data.get("staff"):
        parts.append(f"📎 工作人員：{event_data['staff']}")
    has_meeting = event_data.get("meeting_url") or event_data.get("meeting_id") or event_data.get("meeting_password")
    if has_meeting:
        parts.append("")
        parts.append("💻 線上會議資訊：")
        if event_data.get("meeting_url"):
            parts.append("會議連結：")
            parts.append(event_data["meeting_url"])
        if event_data.get("meeting_id"):
            parts.append(f"會議號：{event_data['meeting_id']}")
        if event_data.get("meeting_password"):
            parts.append(f"會議密碼：{event_data['meeting_password']}")
    if event_data.get("notes"):
        parts.append("")
        parts.append(f"📝 備註：{event_data['notes']}")
    if message_timestamp:
        parts.append("")
        parts.append(f"⏱ 登記時間：{message_timestamp}")

    date_str = event_data["date"]
    start_time = event_data.get("start_time")
    end_time = event_data.get("end_time")
    end_date = event_data.get("end_date")

    event = {
        "summary": event_data.get("title", "未命名行程"),
        "description": "\n".join(parts),
    }

    if start_time:
        # 有具體時間：用 dateTime 格式
        event["start"] = {"dateTime": f"{date_str}T{start_time}:00", "timeZone": "Asia/Taipei"}
        if end_time:
            et_date = end_date if end_date else date_str
            event["end"] = {"dateTime": f"{et_date}T{end_time}:00", "timeZone": "Asia/Taipei"}
        else:
            event["end"] = {"dateTime": f"{date_str}T{start_time}:00", "timeZone": "Asia/Taipei"}
    else:
        # 無具體時間：全天事件，用 date 格式
        event["start"] = {"date": date_str}
        if end_date:
            # Google Calendar 的全天事件 end date 是 exclusive，要加一天
            ed = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            event["end"] = {"date": ed.strftime("%Y-%m-%d")}
        else:
            ed = datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=1)
            event["end"] = {"date": ed.strftime("%Y-%m-%d")}

    if event_data.get("location"):
        event["location"] = event_data["location"]
    return event

def create_calendar_event(event_data, ts=None):
    service = get_calendar_service()
    return service.events().insert(calendarId=GOOGLE_CALENDAR_ID, body=build_event_body(event_data, ts)).execute()

def update_calendar_event(event_id, event_data, ts=None):
    service = get_calendar_service()
    return service.events().update(calendarId=GOOGLE_CALENDAR_ID, eventId=event_id, body=build_event_body(event_data, ts)).execute()

def log_daily_event(event_data):
    global daily_event_log, daily_event_log_date
    now = datetime.now(TW_TZ)
    today_str = now.strftime("%Y-%m-%d")
    if daily_event_log_date != today_str:
        daily_event_log = []
        daily_event_log_date = today_str
    date_str = event_data.get("date", "")
    end_date = event_data.get("end_date")
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        if end_date:
            edt = datetime.strptime(end_date, "%Y-%m-%d")
            meeting_date = f"{dt.month}/{dt.day}-{edt.month}/{edt.day}"
        else:
            meeting_date = f"{dt.month}/{dt.day}"
    except ValueError:
        meeting_date = date_str
    attendees = event_data.get("attendees") or []
    non_staff = [a for a in attendees if not a.startswith("幕僚")]
    daily_event_log.append({
        "meeting_date": meeting_date,
        "staff": event_data.get("staff") or "",
        "location_city": event_data.get("location_city") or "",
        "title": event_data.get("title") or "未命名行程",
        "attendees": "、".join(non_staff) if non_staff else "",
    })

# ===== 請假相關函式 =====

def find_leave_event(service, date_str):
    events_result = service.events().list(
        calendarId=GOOGLE_CALENDAR_ID, timeMin=f"{date_str}T00:00:00+08:00",
        timeMax=f"{date_str}T23:59:59+08:00", q="同仁休假登記", singleEvents=True, orderBy="startTime").execute()
    for evt in events_result.get("items", []):
        if evt.get("summary") == "同仁休假登記":
            return evt
    return None

def parse_existing_leaves(description):
    leaves = []
    if not description:
        return leaves
    for line in description.strip().split("\n"):
        line = line.strip()
        if line and line[0].isdigit() and "." in line:
            leaves.append(line.split(".", 1)[1])
    return leaves

def format_leave_description(entries):
    return "\n".join(f"{i}.{e}" for i, e in enumerate(entries, 1))

def get_leave_name(entry_str):
    leave_types = ["產檢假", "產假", "病假", "喪假", "事假", "補休", "特休", "公假", "婚假", "請休"]
    name_by_type = None
    for lt in leave_types:
        if lt in entry_str:
            name_by_type = entry_str.split(lt)[0].rstrip("0123456789-、 ")
            break
    name_by_digit = ""
    for ch in entry_str:
        if ch.isdigit():
            break
        name_by_digit += ch
    if name_by_type and name_by_digit:
        return name_by_type if len(name_by_type) <= len(name_by_digit) else name_by_digit
    return name_by_type or name_by_digit or entry_str

def parse_segments(entry_str):
    name = get_leave_name(entry_str)
    rest = entry_str[len(name):]
    return name, [s.strip() for s in rest.split("、") if s.strip()]

def parse_time_range(segment):
    m = re.search(r"(\d{3,4})-(\d{3,4})", segment)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def times_overlap(s1, e1, s2, e2):
    if s1 is None or s2 is None:
        return False
    return s1 < e2 and s2 < e1

def format_leave_segment(leave):
    leave_type = leave.get("leave_type") or "請休"
    s = leave.get("start_time", "").replace(":", "")
    e = leave.get("end_time", "").replace(":", "")
    if s and e:
        return f"{s}-{e}{leave_type}"
    else:
        return f"{leave_type}（時間未指定）"

def merge_into_existing(existing_entry, new_segments):
    name, existing_segs = parse_segments(existing_entry)
    result_segs = list(existing_segs)
    for new_seg in new_segments:
        ns, ne = parse_time_range(new_seg)
        result_segs = [
            seg for seg in result_segs
            if not times_overlap(parse_time_range(seg)[0], parse_time_range(seg)[1], ns, ne)
        ]
        result_segs.append(new_seg)
    def sort_key(seg):
        s, _ = parse_time_range(seg)
        return s if s is not None else 9999
    result_segs.sort(key=sort_key)
    return f"{name}{'、'.join(result_segs)}"

def handle_leave(event_data):
    service = get_calendar_service()
    results = []
    leaves_by_date = OrderedDict()
    for leave in event_data.get("leaves", []):
        date_str = leave.get("date")
        if not date_str:
            dates = event_data.get("dates", [])
            date_str = dates[0] if dates else None
        if not date_str:
            continue
        if date_str not in leaves_by_date:
            leaves_by_date[date_str] = []
        leaves_by_date[date_str].append(leave)
    for date_str, day_leaves in leaves_by_date.items():
        existing = find_leave_event(service, date_str)
        new_by_name = OrderedDict()
        for leave in day_leaves:
            name = leave.get("name", "未知")
            seg = format_leave_segment(leave)
            if name not in new_by_name:
                new_by_name[name] = []
            new_by_name[name].append(seg)
        if existing:
            existing_leaves = parse_existing_leaves(existing.get("description", ""))
            for name, new_segs in new_by_name.items():
                found_idx = None
                for i, el in enumerate(existing_leaves):
                    if get_leave_name(el) == name:
                        found_idx = i
                        break
                if found_idx is not None:
                    merged = merge_into_existing(existing_leaves[found_idx], new_segs)
                    existing_leaves[found_idx] = merged
                else:
                    existing_leaves.append(f"{name}{'、'.join(new_segs)}")
            existing["description"] = format_leave_description(existing_leaves)
            service.events().update(calendarId=GOOGLE_CALENDAR_ID, eventId=existing["id"], body=existing).execute()
            results.append({"date": date_str, "action": "updated", "entries": existing_leaves})
        else:
            new_entries = [f"{name}{'、'.join(segs)}" for name, segs in new_by_name.items()]
            end_date = (datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            event = {"summary": "同仁休假登記", "description": format_leave_description(new_entries),
                     "start": {"date": date_str}, "end": {"date": end_date}}
            service.events().insert(calendarId=GOOGLE_CALENDAR_ID, body=event).execute()
            results.append({"date": date_str, "action": "created", "entries": new_entries})
    return results

def reply_message(reply_token, text):
    resp = requests.post("https://api.line.me/v2/bot/message/reply",
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"},
        json={"replyToken": reply_token, "messages": [{"type": "text", "text": text}]})
    if resp.status_code != 200:
        logger.error(f"Line reply failed: {resp.status_code} {resp.text}")

def push_message(user_id, text):
    resp = requests.post("https://api.line.me/v2/bot/message/push",
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"},
        json={"to": user_id, "messages": [{"type": "text", "text": text}]})
    if resp.status_code != 200:
        logger.error(f"Line push failed: {resp.status_code} {resp.text}")

def verify_signature(body, signature):
    h = hmac.new(LINE_CHANNEL_SECRET.encode("utf-8"), body, hashlib.sha256).digest()
    return hmac.compare_digest(base64.b64encode(h).decode("utf-8"), signature)

def format_event_confirmation(event_data, action="登記"):
    lines = [f"✅ 行程已{action}！", "", f"📌 {event_data.get('title', '未命名行程')}"]
    date_str = event_data.get('date', '未指定')
    end_date = event_data.get('end_date')
    if end_date and end_date != date_str:
        lines.append(f"📅 {date_str} ~ {end_date}")
    else:
        lines.append(f"📅 {date_str}")
    s, e = event_data.get("start_time", ""), event_data.get("end_time", "")
    if s and e:
        lines.append(f"🕐 {s} ~ {e}")
    elif s:
        lines.append(f"🕐 {s}")
    if event_data.get("location"):
        lines.append(f"📍 {event_data['location']}")
    if event_data.get("host"):
        lines.append(f"🎤 主持人：{event_data['host']}")
    if event_data.get("attendees"):
        lines.append(f"👥 {'、'.join(event_data['attendees'])}")
    if event_data.get("staff"):
        lines.append(f"📎 工作人員：{event_data['staff']}")
    if event_data.get("meeting_url") or event_data.get("meeting_id") or event_data.get("meeting_password"):
        lines.extend(["", "💻 線上會議資訊："])
        if event_data.get("meeting_url"):
            lines.extend(["會議連結：", event_data["meeting_url"]])
        if event_data.get("meeting_id"):
            lines.append(f"會議號：{event_data['meeting_id']}")
        if event_data.get("meeting_password"):
            lines.append(f"會議密碼：{event_data['meeting_password']}")
    if event_data.get("notes"):
        lines.append(f"📝 {event_data['notes']}")
    return "\n".join(lines)

def format_leave_confirmation(results):
    lines = ["✅ 休假已登記！", ""]
    for r in results:
        lines.append(f"📅 {r['date']}（{'已更新' if r['action'] == 'updated' else '已新增'}）")
        for entry in r["entries"]:
            lines.append(f"  • {entry}")
        lines.append("")
    return "\n".join(lines).strip()

def generate_daily_report_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    for col, h in enumerate(["會議時間", "幕僚", "會議地點", "會議名稱", "出席人員"], 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    if daily_event_log:
        for ri, entry in enumerate(daily_event_log, 2):
            ws.cell(row=ri, column=1, value=entry["meeting_date"])
            ws.cell(row=ri, column=2, value=entry["staff"])
            ws.cell(row=ri, column=3, value=entry["location_city"])
            ws.cell(row=ri, column=4, value=entry["title"])
            ws.cell(row=ri, column=5, value=entry["attendees"])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def send_daily_report():
    if not ADMIN_USER_ID:
        logger.warning("ADMIN_USER_ID not set, skipping daily report")
        return
    global daily_event_log
    now = datetime.now(TW_TZ)
    ds = now.strftime("%m/%d")
    if not daily_event_log:
        push_message(ADMIN_USER_ID, f"📊 今日行程更新報表（{ds}）\n\n今日無新增或修改的行程。")
        daily_event_log = []
        return
    try:
        buf = generate_daily_report_excel()
        fn = f"report_{now.strftime('%Y%m%d')}.xlsx"
        fp = os.path.join(REPORT_DIR, fn)
        with open(fp, "wb") as f:
            f.write(buf.read())
        logger.info(f"Excel report saved: {fp}")
        for old in os.listdir(REPORT_DIR):
            op = os.path.join(REPORT_DIR, old)
            if os.path.isfile(op) and now.timestamp() - os.path.getmtime(op) > 7 * 86400:
                os.remove(op)
        lines = [f"📊 今日行程更新報表（{ds}）", "", f"共 {len(daily_event_log)} 筆行程更新"]
        if NAS_EXTERNAL_URL:
            lines.extend(["", "📥 下載 Excel 報表：", f"{NAS_EXTERNAL_URL}/dl/{fn}"])
        else:
            lines.append("")
            for entry in daily_event_log:
                lines.append(f"• {entry['meeting_date']} | {entry['staff'] or '-'} | {entry['location_city'] or '-'} | {entry['title']} | {entry['attendees'] or '-'}")
        push_message(ADMIN_USER_ID, "\n".join(lines))
    except Exception as e:
        logger.error(f"Daily report error: {e}")
        push_message(ADMIN_USER_ID, f"⚠️ 今日報表產生失敗：{str(e)}")
    daily_event_log = []
    logger.info("Daily report sent, log cleared")

def report_scheduler():
    while True:
        now = datetime.now(TW_TZ)
        target = now.replace(hour=REPORT_HOUR, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        wait = (target - now).total_seconds()
        logger.info(f"Next daily report at {target.strftime('%Y-%m-%d %H:%M')} ({int(wait)}s)")
        time_module.sleep(wait)
        try:
            send_daily_report()
        except Exception as e:
            logger.error(f"Report scheduler error: {e}")

def get_message_timestamp(event):
    ts = event.get("timestamp", 0)
    dt = datetime.fromtimestamp(ts / 1000, tz=TW_TZ) if ts else datetime.now(TW_TZ)
    return dt.strftime("%Y年%m月%d日 %H:%M")

def get_source_id(event):
    src = event.get("source", {})
    return src.get("groupId") or src.get("roomId") or src.get("userId", "unknown")

def cleanup_pending():
    now = datetime.now(TW_TZ)
    for k in [k for k, v in pending_confirmations.items() if now > v.get("expire_time", now)]:
        del pending_confirmations[k]

class WebhookHandler(BaseHTTPRequestHandler):
    def do_POST(self):
        logger.info(f"POST {self.path} from {self.client_address}")
        if self.path != "/callback":
            self.send_response(404)
            self.end_headers()
            return
        body = self.rfile.read(int(self.headers.get("Content-Length", 0)))
        sig = self.headers.get("X-Line-Signature", "")
        if not verify_signature(body, sig):
            self.send_response(403)
            self.end_headers()
            return
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.end_headers()
        self.wfile.write(b'{"status":"ok"}')
        try:
            data = json.loads(body.decode("utf-8"))
            for event in data.get("events", []):
                self.handle_event(event)
        except Exception as e:
            logger.error(f"Error processing webhook: {e}")

    def do_GET(self):
        logger.info(f"GET {self.path} from {self.client_address}")
        decoded_path = unquote(self.path)
        if decoded_path.startswith("/dl/"):
            self.serve_report(decoded_path)
            return
        self.send_response(200)
        self.send_header("Content-Type", "text/plain")
        self.end_headers()
        self.wfile.write("LINE Calendar Bot is running!".encode("utf-8"))

    def serve_report(self, decoded_path):
        filename = decoded_path.split("/dl/", 1)[1]
        if "/" in filename or "\\" in filename or ".." in filename:
            self.send_response(403)
            self.end_headers()
            return
        filepath = os.path.join(REPORT_DIR, filename)
        logger.info(f"Report download: {filename} -> {filepath}")
        if not os.path.exists(filepath):
            self.send_response(404)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write("Report not found".encode("utf-8"))
            return
        with open(filepath, "rb") as f:
            data = f.read()
        logger.info(f"Serving report: {filename} ({len(data)} bytes)")
        self.send_response(200)
        self.send_header("Content-Type", "application/octet-stream")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(data)

    def handle_event(self, event):
        if event.get("type") != "message":
            return
        if event.get("message", {}).get("type") != "text":
            return
        message_text = event["message"]["text"].strip()
        reply_token = event["replyToken"]
        source_id = get_source_id(event)
        cleanup_pending()
        if source_id in pending_confirmations:
            if message_text == "確認":
                self.handle_confirm_replace(source_id, reply_token)
                return
            elif message_text == "新增":
                self.handle_confirm_add(source_id, reply_token)
                return
        if not message_text.startswith("登記"):
            return
        message_text = message_text[2:].strip()
        if not message_text:
            reply_message(reply_token, "請在「登記」後面加上行程資訊，例如：\n登記 3月15日下午兩點，A棟301會議室，產品規劃會議\n\n或請假資訊：\n登記 0213同仁請假 麗如1.5小時(09-1030)\n登記 0213 張三產檢假1小時(09-10)")
            return
        logger.info(f"Received message: {message_text}")
        message_timestamp = get_message_timestamp(event)
        try:
            parsed_data = parse_event_with_claude(message_text)
            logger.info(f"Parsed: {json.dumps(parsed_data, ensure_ascii=False)}")
            msg_type = parsed_data.get("type", "event")
            if msg_type == "leave":
                if not parsed_data.get("leaves"):
                    reply_message(reply_token, "⚠️ 無法從訊息中辨識出請假資訊，請確認訊息中有包含日期和請假人員。")
                    return
                results = handle_leave(parsed_data)
                reply_message(reply_token, format_leave_confirmation(results))
            else:
                if not parsed_data.get("date"):
                    reply_message(reply_token, "⚠️ 無法從訊息中辨識出日期，請確認訊息中有包含行程的日期。")
                    return
                service = get_calendar_service()
                duplicate = find_duplicate_event(service, parsed_data)
                if duplicate:
                    pending_confirmations[source_id] = {
                        "event_data": parsed_data, "existing_event_id": duplicate["id"],
                        "message_timestamp": message_timestamp,
                        "expire_time": datetime.now(TW_TZ) + timedelta(minutes=10),
                    }
                    dup_title = duplicate.get("summary", "未命名")
                    dup_start, dup_end = "", ""
                    try:
                        dup_start = datetime.fromisoformat(duplicate["start"]["dateTime"]).strftime("%H:%M")
                    except Exception:
                        pass
                    try:
                        dup_end = datetime.fromisoformat(duplicate["end"]["dateTime"]).strftime("%H:%M")
                    except Exception:
                        pass
                    time_str = f"{dup_start}~{dup_end}" if dup_start and dup_end else ""
                    reply_message(reply_token, f"⚠️ 發現同一天已有類似行程：\n📌 {dup_title}\n🕐 {time_str}\n\n請回覆：\n「確認」→ 替換舊行程\n「新增」→ 另外新增一筆")
                    return
                created = create_calendar_event(parsed_data, message_timestamp)
                logger.info(f"Event created: {created.get('id')}")
                log_daily_event(parsed_data)
                reply_message(reply_token, format_event_confirmation(parsed_data))
        except json.JSONDecodeError as e:
            logger.error(f"JSON parse error: {e}")
            reply_message(reply_token, "⚠️ 解析行程資訊時發生錯誤，請確認訊息格式是否包含行程相關資訊。")
        except Exception as e:
            logger.error(f"Error: {e}")
            reply_message(reply_token, f"⚠️ 發生錯誤：{str(e)}")

    def handle_confirm_replace(self, source_id, reply_token):
        pending = pending_confirmations.pop(source_id, None)
        if not pending:
            reply_message(reply_token, "⚠️ 找不到待確認的行程，請重新登記。")
            return
        try:
            ed = pending["event_data"]
            updated = update_calendar_event(pending["existing_event_id"], ed, pending["message_timestamp"])
            logger.info(f"Event replaced: {updated.get('id')}")
            log_daily_event(ed)
            reply_message(reply_token, format_event_confirmation(ed, action="替換"))
        except Exception as e:
            logger.error(f"Replace error: {e}")
            reply_message(reply_token, f"⚠️ 替換行程時發生錯誤：{str(e)}")

    def handle_confirm_add(self, source_id, reply_token):
        pending = pending_confirmations.pop(source_id, None)
        if not pending:
            reply_message(reply_token, "⚠️ 找不到待確認的行程，請重新登記。")
            return
        try:
            ed = pending["event_data"]
            created = create_calendar_event(ed, pending["message_timestamp"])
            logger.info(f"Event added: {created.get('id')}")
            log_daily_event(ed)
            reply_message(reply_token, format_event_confirmation(ed))
        except Exception as e:
            logger.error(f"Add error: {e}")
            reply_message(reply_token, f"⚠️ 新增行程時發生錯誤：{str(e)}")

    def log_message(self, format, *args):
        pass

if __name__ == "__main__":
    logger.info(f"Starting LINE Calendar Bot on port {PORT}...")
    threading.Thread(target=report_scheduler, daemon=True).start()
    logger.info(f"Daily report scheduler started (send at {REPORT_HOUR}:00)")
    server = HTTPServer(("0.0.0.0", PORT), WebhookHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        logger.info("Shutting down...")
        server.server_close()
